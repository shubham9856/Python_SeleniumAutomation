# from selenium.webdriver.chrome import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

file_path = "/MockData/Xcel_Mockdata.xlsx"
fruit_name = "Apple"
newValue = "990"


def update_excel_data(filePath, searchTerm, colName, new_value):
    """
    Method to find the specific data in the Excel
    :param filePath: Defines the excel sheet file path
    :param searchTerm: defines the term/ data to find in the cells
    :param colName: defines the expected column name to search the data/ term
    :param new_value: defines the data to be updated
    :return:
    """
    # Create object of the file
    book = openpyxl.load_workbook(filePath)
    # Get the active sheet in Excel
    sheet = book.active
    # Dictionary to hold the row and col number to update the data in specific cell
    Dict = {}

    # This loop will iterate all cells in first row
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    # This loop will iterate each cell of excel
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    # save the changes into Excel
    book.save(file_path)


def update_excel_sheet_data(driver_instance):
    """
    This method searches a specific data in the excel sheet and updates it with the specified value
    :param driver_instance: Driver object is taken from conftest file
    """
    driver = driver_instance
    driver.implicitly_wait(5)
    driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
    driver.find_element(By.ID, "downloadButton").click()

    # edit the Excel with updated value
    update_excel_data(file_path, fruit_name, "price", newValue)

    # upload
    file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
    file_input.send_keys(file_path)

    wait = WebDriverWait(driver, 5)
    toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
    wait.until(expected_conditions.visibility_of_element_located(toast_locator))
    print(driver.find_element(*toast_locator).text)
    priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
    actual_price = driver.find_element(By.XPATH,
                                       "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']").text
    assert actual_price == newValue
